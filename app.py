import streamlit as st
import openpyxl
import io
import re
import gspread
from PIL import Image as PILImage
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

# --- CONFIGURAÇÕES DA PÁGINA ---
st.set_page_config(page_title="Extrator de Agulhas", page_icon="🪡")
st.title("🪡 Portal de Extração de Relatórios")

# --- AUTENTICAÇÃO VIA SECRETS ---
def get_gspread_client():
    # Carrega as credenciais das 'Secrets' do Streamlit
    creds_dict = st.secrets["gcp_service_account"]
    creds = service_account.Credentials.from_service_account_info(
        creds_dict, 
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
    )
    return gspread.authorize(creds), creds

# --- FUNÇÃO PARA CRIAR PASTAS NO DRIVE ---
def get_or_create_folder(drive_service, folder_name, parent_id):
    query = f"name = '{folder_name}' and '{parent_id}' in parents and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    response = drive_service.files().list(q=query).execute()
    folders = response.get('files', [])
    if folders:
        return folders[0]['id']
    else:
        file_metadata = {
            'name': folder_name,
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [parent_id]
        }
        folder = drive_service.files().create(body=file_metadata, fields='id').execute()
        return folder.get('id')

# --- CONFIGURAÇÕES DE MAPEAMENTO ---
ID_PLANILHA = st.secrets["spreadsheet_id"]
ID_PASTA_RAIZ = st.secrets["drive_folder_id"]
NOME_ABA = "Página1"

# Excel : Google (0=A, 1=B...)
mapeamento = {4: 0, 1: 1, 2: 2, 29: 3} 
total_cols_google = 5

meses_nomes = {
    "01": "01 - Janeiro", "02": "02 - Fevereiro", "03": "03 - Março",
    "04": "04 - Abril", "05": "05 - Maio", "06": "06 - Junho",
    "07": "07 - Julho", "08": "08 - Agosto", "09": "09 - Setembro",
    "10": "10 - Outubro", "11": "11 - Novembro", "12": "12 - Dezembro"
}

# --- INTERFACE ---
uploaded_file = st.file_uploader("Arraste o relatório .xlsx aqui", type="xlsx")

if uploaded_file:
    if st.button("🚀 Processar e Enviar para o Google"):
        try:
            client, creds = get_gspread_client()
            drive_service = build('drive', 'v3', credentials=creds)
            sheet_google = client.open_by_key(ID_PLANILHA).worksheet(NOME_ABA)
            
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            ws = wb.active
            
            dados_para_google = []
            
            with st.status("Extraindo imagens e dados...", expanded=True) as status:
                if hasattr(ws, '_images'):
                    for img in ws._images:
                        col = img.anchor._from.col + 1
                        row = img.anchor._from.row + 1
                        
                        if col == 29: # Coluna AC
                            val_ac = str(ws.cell(row=row, column=29).value or "")
                            match = re.search(r'(\d{4})-(\d{2})[\\/]([^\\/]+\.[pP][nN][gG])', val_ac)
                            
                            if match:
                                ano, mes_num, nome_foto = match.groups()
                                
                                # 1. Organizar pastas no Drive
                                id_ano = get_or_create_folder(drive_service, ano, ID_PASTA_RAIZ)
                                id_mes = get_or_create_folder(drive_service, meses_nomes.get(mes_num, mes_num), id_ano)
                                
                                # 2. Upload da Imagem para o Drive
                                try:
                                    img_bytes = img.ref.open().read()
                                except:
                                    img.ref.seek(0)
                                    img_bytes = img.ref.read()
                                
                                file_metadata = {'name': nome_foto, 'parents': [id_mes]}
                                media = MediaIoBaseUpload(io.BytesIO(img_bytes), mimetype='image/png')
                                drive_service.files().create(body=file_metadata, media_body=media).execute()
                                
                                # 3. Coletar dados da linha
                                nova_linha = [""] * total_cols_google
                                for ex_col, go_idx in mapeamento.items():
                                    nova_linha[go_idx] = str(ws.cell(row=row, column=ex_col).value or "")
                                nova_linha[-1] = f"{ano}/{mes_num}/{nome_foto}"
                                
                                dados_para_google.append(nova_linha)
                                st.write(f"✅ Linha {row} processada.")

                if dados_para_google:
                    sheet_google.append_rows(dados_para_google)
                    status.update(label="Processamento concluído!", state="complete")
                    st.success(f"{len(dados_para_google)} registros enviados!")
                    st.balloons()
                else:
                    st.warning("Nenhuma imagem válida encontrada na coluna AC.")
                    
        except Exception as e:
            st.error(f"Ocorreu um erro: {e}")