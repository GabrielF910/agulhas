import openpyxl
import os
import io
import re
import gspread
from google.auth import default
from PIL import Image as PILImage
from google.colab import files, drive, auth

# 1. AUTENTICAÇÃO
auth.authenticate_user()
creds, _ = default()
gc = gspread.authorize(creds)
drive.mount('/content/drive')

# 2. CONFIGURAÇÕES DE MAPEAMENTO (AJUSTE AQUI)
# ---------------------------------------------------------------------
id_planilha_destino = 'SEU_ID_DA_PLANILHA_AQUI'
nome_aba_destino = 'Página1'
nome_do_atalho_drive = 'NOME_DO_ATALHO_AQUI'

# MAPEAMENTO: { Coluna_Excel : Posição_na_Lista_do_Google }
# Se você quer que a Coluna D (4) do Excel caia na Coluna A (1) do Google:
# O primeiro item da lista (índice 0) é a Coluna A do Google, o segundo (1) é a B, etc.
mapeamento = {
    4:  0,  # Coluna D (Excel) -> Coluna A (Google)
    1:  1,  # Coluna A (Excel) -> Coluna B (Google)
    2:  2,  # Coluna B (Excel) -> Coluna C (Google)
    29: 3   # Coluna AC (Excel) -> Coluna D (Google)
}
# Quantidade total de colunas que você quer preencher no Google Sheets
total_colunas_google = 5 
# ---------------------------------------------------------------------

caminho_base = f'/content/drive/MyDrive/{nome_do_atalho_drive}'
planilha_google = gc.open_by_key(id_planilha_destino)
aba_destino = planilha_google.worksheet(nome_aba_destino)

meses_nomes = {
    "01": "01 - Janeiro", "02": "02 - Fevereiro", "03": "03 - Março",
    "04": "04 - Abril", "05": "05 - Maio", "06": "06 - Junho",
    "07": "07 - Julho", "08": "08 - Agosto", "09": "09 - Setembro",
    "10": "10 - Outubro", "11": "11 - Novembro", "12": "12 - Dezembro"
}

# 3. UPLOAD E PROCESSAMENTO
uploaded = files.upload()
if uploaded:
    nome_arquivo = list(uploaded.keys())[0]
    wb = openpyxl.load_workbook(nome_arquivo, data_only=True)
    sheet = wb.active
    
    dados_finais = []
    
    if hasattr(sheet, '_images'):
        for image in sheet._images:
            try:
                row_ancora = image.anchor._from.row + 1
                col_ancora = image.anchor._from.col + 1
                
                # Filtro: Apenas coluna AC (29)
                if col_ancora == 29:
                    valor_ac = sheet.cell(row=row_ancora, column=29).value
                    match = re.search(r'(\d{4})-(\d{2})[\\/]([^\\/]+\.[pP][nN][gG])', str(valor_ac))
                    
                    if match:
                        ano, mes_num, nome_foto = match.groups()
                        
                        # --- SALVAR FOTO NO DRIVE ---
                        pasta_destino = os.path.join(caminho_base, ano, meses_nomes.get(mes_num, mes_num))
                        if not os.path.exists(pasta_destino): os.makedirs(pasta_destino)
                        
                        # Extração dos bytes da imagem
                        try:
                            img_data = image.ref.open().read()
                        except AttributeError:
                            image.ref.seek(0)
                            img_data = image.ref.read()
                        
                        PILImage.open(io.BytesIO(img_data)).save(os.path.join(pasta_destino, nome_foto))

                        # --- CONSTRUIR A LINHA DO GOOGLE SHEETS ---
                        # Criamos uma linha vazia com o tamanho total desejado
                        nova_linha = [""] * total_colunas_google
                        
                        for col_excel, pos_google in mapeamento.items():
                            valor = sheet.cell(row=row_ancora, column=col_excel).value
                            nova_linha[pos_google] = str(valor) if valor is not None else ""
                        
                        # Exemplo: Colocar o link/caminho da foto na última coluna manualmente
                        nova_linha[-1] = f"{ano}/{mes_num}/{nome_foto}"
                        
                        dados_finais.append(nova_linha)
                        print(f"✅ Linha {row_ancora} preparada.")

            except Exception as e:
                print(f"❌ Erro na linha {row_ancora}: {e}")

        # 4. ENVIAR PARA O GOOGLE
        if dados_finais:
            aba_destino.append_rows(dados_finais)
            print(f"\n🚀 {len(dados_finais)} linhas enviadas com sucesso!")